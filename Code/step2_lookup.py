"""
步骤二：从C表/D表查找并填充 B表(Pooling表)
=============================================
匹配规则:
  B列以 HGC-Lib- / HGC-POOL- 开头 → C表 (搜所有sheet, F列匹配)
  其他 → D表 (D列匹配)

HGC类型 (C表 → B表):
  L列(Qubit浓度) → C列 + N列
  V列(文库结构)  → O列
  W列(磷酸化*)   → P列
  X列(环化*)     → Q列
  O列(板号)      → K列
  J列(孔号)      → 不填

非HGC类型 (D表 → B表):
  K列(Qubit浓度) → C列 + N列
  O列(结果评价)  → O列
  T列(版号)      → K列
  S列(孔位)      → J列
"""
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

from pathlib import Path
BASE = Path(__file__).resolve().parent.parent / 'excel_data'
SRC_C = str(BASE / '外包文库质检总表2026(1).xlsx')
SRC_D = str(BASE / '2026年1月-12月自建库出库报告-T7.xlsx')
DST   = str(BASE / '20260711文库pooling表T7+PE150-zss.xlsx')
SHEETS = ['A', 'B', 'C']
CENTER = Alignment(horizontal='center', vertical='center')

# ── B表目标列 (1-based) ──
B_C = 3     # Qubit浓度
B_D = 4     # 数据量
B_J = 10    # 孔号
B_K = 11    # 板号
B_M = 13    # 平均片段
B_N = 14    # Qubit浓度 (第二处)
B_O = 15    # 文库结构
B_P = 16    # 磷酸化*
B_Q = 17    # 环化*


def build_c_lookup():
    """构建C表查找字典: {HGC编号: row_data}，搜所有sheet"""
    wb = openpyxl.load_workbook(SRC_C, data_only=True)
    lookup = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        # 找表头行，确定列位置
        header_row = None
        col_f = col_l = col_o = col_v = col_w = col_x = col_m = None
        for row in range(1, min(5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                v = str(ws.cell(row=row, column=col).value or '')
                if 'HGC编号' in v:
                    col_f = col
                if 'Qubit浓度' in v:
                    col_l = col
                if v == '板号':
                    col_o = col
                if '文库结构' in v:
                    col_v = col
                if '磷酸化' in v:
                    col_w = col
                if '环化' in v:
                    col_x = col
                if '平均片段' in v:
                    col_m = col
            if col_f:
                header_row = row
                break

        if not col_f:
            continue

        print(f'  C表[{sn}]: header_row={header_row}, F={col_f}, L={col_l}, O={col_o}, V={col_v}, W={col_w}, X={col_x}, M={col_m}')

        for row in range(header_row + 1, ws.max_row + 1):
            key = ws.cell(row=row, column=col_f).value
            if key is None:
                continue
            key = str(key).strip()
            if not key:
                continue

            def get_val(c):
                return ws.cell(row=row, column=c).value if c else None

            lookup[key] = {
                'qubit':  get_val(col_l),  # Qubit浓度
                'plate':  get_val(col_o),  # 板号
                'struct': get_val(col_v),  # 文库结构
                'phos':   get_val(col_w),  # 磷酸化*
                'circ':   get_val(col_x),  # 环化*
                'frag':   get_val(col_m),  # 平均片段
            }

    wb.close()
    print(f'  C表共加载 {len(lookup)} 条记录')
    return lookup


def build_d_lookup():
    """构建D表查找字典: {样本编号: row_data}"""
    wb = openpyxl.load_workbook(SRC_D, data_only=True)
    ws = wb['自建库出库报告总表']

    col_d = col_k = col_o = col_s = col_t = col_m = None
    for col in range(1, ws.max_column + 1):
        v = str(ws.cell(row=1, column=col).value or '')
        if v == '样本编号':
            col_d = col
        elif 'Qubit浓度' in v:
            col_k = col
        elif '结果评价' in v:
            col_o = col
        elif v == '孔位':
            col_s = col
        elif v == '版号':
            col_t = col
        elif '平均片段' in v:
            col_m = col

    print(f'  D表: D={col_d}, K={col_k}, O={col_o}, S={col_s}, T={col_t}, M={col_m}')

    lookup = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=col_d).value
        if key is None:
            continue
        key = str(key).strip()
        if not key:
            continue

        def get_val(c):
            return ws.cell(row=row, column=c).value if c else None

        lookup[key] = {
            'qubit':  get_val(col_k),
            'eval':   get_val(col_o),
            'hole':   get_val(col_s),
            'plate':  get_val(col_t),
            'frag':   get_val(col_m),
        }

    wb.close()
    print(f'  D表共加载 {len(lookup)} 条记录')
    return lookup


def is_hgc_type(lib_id):
    """判断是否为HGC类型"""
    if lib_id is None:
        return False
    s = str(lib_id).strip()
    return s.startswith('HGC-Lib-') or s.startswith('HGC-POOL-')


def fill_sheet(ws_b, c_lookup, d_lookup, name):
    """填充单个工作表"""
    filled = 0
    missed = 0

    for row in range(2, ws_b.max_row + 1):
        lib_id = ws_b.cell(row=row, column=2).value
        if lib_id is None:
            continue
        lib_id = str(lib_id).strip()

        if is_hgc_type(lib_id):
            data = c_lookup.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_N, data.get('qubit'))
                write_cell(ws_b, row, B_O, data.get('struct'))
                write_cell(ws_b, row, B_P, data.get('phos'))
                write_cell(ws_b, row, B_Q, data.get('circ'))
                write_cell(ws_b, row, B_K, data.get('plate'))
                write_cell(ws_b, row, B_M, data.get('frag'))
                filled += 1
            else:
                missed += 1
        else:
            data = d_lookup.get(lib_id)
            if data:
                write_cell(ws_b, row, B_C, data.get('qubit'))
                write_cell(ws_b, row, B_N, data.get('qubit'))
                write_cell(ws_b, row, B_O, data.get('eval'))
                write_cell(ws_b, row, B_K, data.get('plate'))
                write_cell(ws_b, row, B_J, data.get('hole'))
                write_cell(ws_b, row, B_M, data.get('frag'))
                filled += 1
            else:
                missed += 1

    print(f'  Sheet {name}: 匹配={filled}, 未匹配={missed}')


def write_cell(ws, row, col, value):
    """写入单元格并设置居中"""
    if col is None:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = CENTER


def main():
    print('构建C表查找字典...')
    c_lookup = build_c_lookup()

    print('构建D表查找字典...')
    d_lookup = build_d_lookup()

    wb_dst = openpyxl.load_workbook(DST)

    for name in SHEETS:
        print(f'\n处理工作表 [{name}]')
        fill_sheet(wb_dst[name], c_lookup, d_lookup, name)

    wb_dst.save(DST)
    print(f'\n{"="*50}\n✅ 步骤二完成 → {DST}')


if __name__ == '__main__':
    main()
