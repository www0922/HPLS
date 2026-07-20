"""
清空 Pooling 表内容，只保留表头
"""
from config import DST
import openpyxl

POOLING = DST

wb = openpyxl.load_workbook(POOLING)

for sn in wb.sheetnames:
    ws = wb[sn]

    # 解除第2行起的所有合并单元格（表头行1不动）
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 2:
            ws.unmerge_cells(str(mr))

    # 删除第2行起的所有数据行（表头行1不动）
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    print(f'Sheet [{sn}]: 已清空')

wb.save(POOLING)
print(f'\n✅ 完成 → {POOLING}')
