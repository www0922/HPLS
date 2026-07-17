"""
步骤六：T7+制备 数据填充与公式
================================
1. 从文库环化传递数据到 T7+制备
2. A列 laneID = 当天日期 + Lane首字母 (如 0717A)
3. 相同 laneID 为一组, 组间空2行
4. 填入公式: F,G,I,K,L,M,Q
5. 每组汇总行: H_sum, K_sum, M_sum
"""
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side

BASE = Path(__file__).resolve().parent.parent / 'excel_data'
DST = str(BASE / '20260711文库pooling表T7+PE150-zss.xlsx')
CENTER = Alignment(horizontal='center', vertical='center')
ALL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# T7+制备 列索引
COL_A = 1    # laneID
COL_B = 2    # 文库ssDNA编号
COL_C = 3    # SSDNA浓度 (用户填)
COL_D = 4    # 平均片段 (用户填)
COL_E = 5    # 偏差判断
COL_F = 6    # 150fmol对应质量
COL_G = 7    # 投入摩尔数
COL_H = 8    # 数据量
COL_I = 9    # 理论数据量占比
COL_J = 10   # 调整比例 (用户填)
COL_K = 11   # K列
COL_L = 12   # 实际投入占比
COL_M = 13   # ssDNA取样体积
COL_N = 14   # 文库类型
COL_O = 15   # 客户单位
COL_P = 16   # TE补充体积
COL_Q = 17   # 实际调整比例


def fill_t7_sheet(ws_t7, hh_rows):
    """填充 T7+制备 工作表"""
    # 清空数据行
    if ws_t7.max_row > 1:
        ws_t7.delete_rows(2, ws_t7.max_row - 1)

    # 当天日期 MMDD 格式
    date_str = datetime.now().strftime('%m%d')

    # 按 laneID 首字母分组
    groups = {}  # {letter: [row_data, ...]}
    for rd in hh_rows:
        letter = rd['lane'][0]  # A1→A, B2→B, C3→C
        if letter not in groups:
            groups[letter] = []
        groups[letter].append(rd)

    current_row = 2

    for letter in sorted(groups.keys()):
        lane_id = f'{date_str}{letter}'
        rows = groups[letter]
        group_start = current_row
        group_end = current_row + len(rows) - 1

        for rd in rows:
            # B列: 文库ssDNA编号
            c = ws_t7.cell(row=current_row, column=COL_A)
            c.value = lane_id
            c.alignment = CENTER

            # B列: 文库ssDNA编号
            c = ws_t7.cell(row=current_row, column=COL_B)
            c.value = rd['lib_id']
            c.alignment = CENTER

            # H列: 数据量
            c = ws_t7.cell(row=current_row, column=COL_H)
            c.value = rd['data_g']
            c.alignment = CENTER

            # C, D, J 留空(用户填)

            # E列: D与中位值比较
            c = ws_t7.cell(row=current_row, column=COL_E)
            c.value = f'=IF(ABS(D{current_row}-D{group_end + 1})<=100,"正常",IF(D{current_row}>D{group_end + 1},"偏大"&INT(D{current_row}-D{group_end + 1})&"bp","偏小"&INT(D{group_end + 1}-D{current_row})&"bp"))'
            c.alignment = CENTER

            # F列: =D*0.33*G*0.001
            c = ws_t7.cell(row=current_row, column=COL_F)
            c.value = f'=D{current_row}*0.33*G{current_row}*0.001'
            c.alignment = CENTER

            # G列: =900*L
            c = ws_t7.cell(row=current_row, column=COL_G)
            c.value = f'=900*L{current_row}'
            c.alignment = CENTER

            # I列: =ROUND(H / H_summary, 4)
            c = ws_t7.cell(row=current_row, column=COL_I)
            c.value = f'=ROUND(H{current_row}/H{group_end + 1},4)'
            c.alignment = CENTER

            # K列: =I*J
            c = ws_t7.cell(row=current_row, column=COL_K)
            c.value = f'=I{current_row}*J{current_row}'
            c.alignment = CENTER

            # L列: =K / K_summary
            c = ws_t7.cell(row=current_row, column=COL_L)
            c.value = f'=K{current_row}/K{group_end + 1}'
            c.alignment = CENTER

            # M列: =F/C
            c = ws_t7.cell(row=current_row, column=COL_M)
            c.value = f'=F{current_row}/C{current_row}'
            c.alignment = CENTER

            # N列: 文库类型
            c = ws_t7.cell(row=current_row, column=COL_N)
            c.value = rd['lib_type']
            c.alignment = CENTER

            # O列: 客户单位
            c = ws_t7.cell(row=current_row, column=COL_O)
            c.value = rd['customer']
            c.alignment = CENTER

            # Q列: =I*J
            c = ws_t7.cell(row=current_row, column=COL_Q)
            c.value = f'=I{current_row}*J{current_row}'
            c.alignment = CENTER

            current_row += 1

        # 汇总行: D列=MEDIAN, H_sum, K_sum, M_sum, P列
        summary_row = current_row
        c = ws_t7.cell(row=summary_row, column=COL_D)
        c.value = f'=MEDIAN(D{group_start}:D{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_H)
        c.value = f'=SUM(H{group_start}:H{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_K)
        c.value = f'=SUM(K{group_start}:K{group_end})'
        c.alignment = CENTER

        c = ws_t7.cell(row=summary_row, column=COL_M)
        c.value = f'=SUM(M{group_start}:M{group_end})'
        c.alignment = CENTER

        # P列: =96 - M汇总
        c = ws_t7.cell(row=summary_row, column=COL_P)
        c.value = f'=96-M{summary_row}'
        c.alignment = CENTER

        # 仅数据行加框线
        for r in range(group_start, group_end + 1):
            for col in range(1, 18):
                ws_t7.cell(row=r, column=col).border = ALL_BORDER

        current_row += 1

    # 数据行行高设为30
    for row in range(2, current_row):
        ws_t7.row_dimensions[row].height = 30

    print(f'  T7+制备: {len(groups)} 组, 共 {current_row - 2} 行')


def main():
    wb = openpyxl.load_workbook(DST)

    # 读取文库环化数据
    ws_hh = wb['文库环化']
    hh_rows = []
    for row in range(2, ws_hh.max_row + 1):
        lane = ws_hh.cell(row=row, column=1).value  # A列
        lib_id = ws_hh.cell(row=row, column=1).value  # A列=文库编号
        data_g = ws_hh.cell(row=row, column=5).value  # E列=数据量
        lib_type = ws_hh.cell(row=row, column=8).value  # H列=文库类型
        customer = ws_hh.cell(row=row, column=9).value  # I列=客户单位
        if lane:
            hh_rows.append({
                'lane': str(lane).strip(),
                'lib_id': lib_id,
                'data_g': data_g,
                'lib_type': lib_type,
                'customer': customer,
            })

    print(f'文库环化: {len(hh_rows)} 行')

    # 填充 T7+制备
    print(f'\n{"─"*50}')
    print('填充 [T7+制备]')
    print(f'{"─"*50}')
    fill_t7_sheet(wb['T7+制备'], hh_rows)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'✅ 步骤六完成 → {DST}')


if __name__ == '__main__':
    main()
