"""
步骤四：B表 公式计算、重分组、体积调整、汇总
=============================================
对B表(20260711文库pooling表T7+PE150-zss.xlsx)的A/B/C工作表:

任务:
  1. 计算 E列(文库摩尔质量)、F列(最终取样体积)
  2. 检查重分组条件(总数据量>1000G 或 max/min>5)
  3. 重分组后组内按K列(板号)排序
  4. 体积调整(F<0.250 → 放大)
  5. 填写汇总行
"""
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from config import DST
SHEETS = ['A', 'B', 'C']
CENTER = Alignment(horizontal='center', vertical='center')
LIGHT_BLUE = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')

# ── 列索引 (1-based) ──
COL_A = 1    # Lane编号
COL_B = 2    # 文库编号
COL_C = 3    # Qubit浓度
COL_D = 4    # 数据量
COL_E = 5    # 文库摩尔质量
COL_F = 6    # 最终取样体积
COL_G = 7    # 文库类型
COL_H = 8    # 客户单位
COL_I = 9    # TE Buffer补充体积
COL_J = 10   # 孔号
COL_K = 11   # 板号
COL_L = 12   # 备注
COL_N = 14   # Qubit浓度(二)
COL_O = 15   # 文库结构
COL_P = 16   # 磷酸化*
COL_M = 13   # 平均片段
COL_Q = 17   # 环化*


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0



def read_groups(ws):
    """读取工作表数据行, 按空白行分隔为组"""
    max_col = ws.max_column
    groups = []
    current = []

    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in values):
            if current:
                groups.append(current)
                current = []
            continue

        rd = {
            'B': ws.cell(row=row, column=COL_B).value,
            'C': safe_float(ws.cell(row=row, column=COL_C).value),
            'D': safe_float(ws.cell(row=row, column=COL_D).value),
            'N': safe_float(ws.cell(row=row, column=COL_N).value),
            'G': ws.cell(row=row, column=COL_G).value,
            'K': ws.cell(row=row, column=COL_K).value,
            'O': ws.cell(row=row, column=COL_O).value,
            'P': ws.cell(row=row, column=COL_P).value,
            'category': str(ws.cell(row=row, column=COL_E).value or ''),  # 步骤三写入的分类标记
            'cells': {c: ws.cell(row=row, column=c).value for c in range(1, max_col + 1)},
        }
        current.append(rd)

    if current:
        groups.append(current)
    return groups


def calc_EF(rows, molar_mass):
    """计算组内每行的 E列 和 F列
    转化文库: N列(Qubit) < 2.5 → ×10, 否则 ×50
    直接环化: 固定 ×300
    """
    d_sum = sum(r['D'] for r in rows)
    if d_sum == 0:
        for r in rows:
            r['E'] = 0.0
            r['F'] = 0.0
        return

    is_conv = (molar_mass == 50.0)
    for r in rows:
        if is_conv:
            n_val = safe_float(r.get('N'))
            mm = 10.0 if n_val < 2.5 else 50.0
        else:
            mm = molar_mass
        r['E'] = round(r['D'] / d_sum * mm, 3)
        r['F'] = round(r['E'] / r['C'], 3) if r['C'] > 0 else 0.0


def regroup(rows, molar_mass):
    """
    检查重分组条件, 必要时拆分并重新计算
    返回: [(sub_rows, sub_molar_mass), ...]
    """
    calc_EF(rows, molar_mass)
    d_sum = sum(r['D'] for r in rows)
    f_vals = [r['F'] for r in rows if r['F'] > 0]

    if not f_vals:
        return [(rows, molar_mass)]

    need = (d_sum > 1000) or (max(f_vals) / min(f_vals) > 5 if min(f_vals) > 0 else False)
    if not need:
        sort_by_plate(rows)
        return [(rows, molar_mass)]

    # ── 重分组 ──
    sorted_by_f = sorted(rows, key=lambda r: r['F'])

    sub_groups = []
    cur = [sorted_by_f[0]]
    cur_min = cur_max = sorted_by_f[0]['F']

    for r in sorted_by_f[1:]:
        new_min = min(cur_min, r['F'])
        new_max = max(cur_max, r['F'])
        if new_min > 0 and new_max / new_min > 5:
            sub_groups.append(cur)
            cur = [r]
            cur_min = cur_max = r['F']
        else:
            cur.append(r)
            cur_min = new_min
            cur_max = new_max

    if cur:
        sub_groups.append(cur)

    # 每个子组: B列升序 → 重新计算, 再板号排序
    result = []
    for sg in sub_groups:
        sg.sort(key=lambda r: str(r['B'] or ''))
        calc_EF(sg, molar_mass)
        sort_by_plate(sg)
        result.append((sg, molar_mass))

    return result


def sort_by_plate(rows):
    """组内按K列(板号)排序, 无板号排末尾"""
    def key(r):
        k = r.get('K')
        return (0, str(k)) if (k is not None and str(k).strip()) else (1, '')
    rows.sort(key=key)


def apply_scale(rows):
    """
    检查F列最小值, 若 <0.250 则找最小整数N使 min(F*N)>0.500
    返回: scale_factor (1 表示未放大)
    """
    f_vals = [r['F'] for r in rows if r['F'] > 0]
    if not f_vals or min(f_vals) >= 0.250:
        return 1

    n = 1
    while min(f_vals) * n <= 0.500:
        n += 1

    for r in rows:
        r['F'] = round(r['F'] * n, 3)
    return n


def process_sheet(ws, name):
    """处理单个工作表"""
    max_col = ws.max_column
    groups = read_groups(ws)
    print(f'  Sheet {name}: 读取 {len(groups)} 组')

    # ── 处理每个组 ──
    all_final = []  # [(rows, molar_mass, scale, is_direct), ...]

    for g in groups:
        # 从步骤三的标记读取分类
        is_direct = (g[0].get('category', '') == '直接环化文库') if g else False
        molar_mass = 300.0 if is_direct else 50.0

        # 重分组 → [(sub_rows, sub_molar_mass), ...]
        sub_results = regroup(g, molar_mass)

        for sg_rows, sg_mm in sub_results:
            scale = apply_scale(sg_rows)
            all_final.append((sg_rows, sg_mm, scale, is_direct))

    regroup_count = len(all_final) - len(groups)
    print(f'    处理: {len(groups)} → {len(all_final)} 组 (新增 {regroup_count})')

    # ── 清空并写回 ──
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    current_row = 2
    for sg_rows, sg_mm, scale, is_direct in all_final:
        d_first = current_row
        d_last = current_row + len(sg_rows) - 1
        summary_row = d_last + 1

        # 写入数据行
        for rd in sg_rows:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col)
                if col == COL_E:
                    cell.value = rd['E']
                elif col == COL_F:
                    cell.value = f'=ROUND(E{current_row}/C{current_row}*G{summary_row},3)'
                else:
                    cell.value = rd['cells'].get(col)
                cell.alignment = CENTER
            current_row += 1

        # ── 汇总行 ──
        # D列: =ROUND(SUM,2)
        c = ws.cell(row=summary_row, column=COL_D)
        c.value = f'=ROUND(SUM(D{d_first}:D{d_last}),2)'
        c.alignment = CENTER

        # E列: 总摩尔质量(静态)
        c = ws.cell(row=summary_row, column=COL_E)
        c.value = round(sum(r['E'] for r in sg_rows), 3)
        c.alignment = CENTER

        # F列: =IF(G>1,ROUND(SUM/G*2,3),ROUND(SUM,3))
        c = ws.cell(row=summary_row, column=COL_F)
        c.value = f'=IF(G{summary_row}>1,ROUND(SUM(F{d_first}:F{d_last})/G{summary_row}*2,3),ROUND(SUM(F{d_first}:F{d_last}),3))'
        c.alignment = CENTER

        # G列: 倍数(静态)
        c = ws.cell(row=summary_row, column=COL_G)
        c.value = scale
        c.alignment = CENTER

        # I列: =ROUND(20/40-F,3), 浅蓝背景
        c = ws.cell(row=summary_row, column=COL_I)
        c.value = f'=ROUND(20-F{summary_row},3)' if not is_direct else f'=ROUND(40-F{summary_row},3)'
        c.alignment = CENTER
        c.fill = LIGHT_BLUE

        # M列: =AVERAGE(片段范围)
        c = ws.cell(row=summary_row, column=COL_M)
        c.value = f'=AVERAGE(M{d_first}:M{d_last})'
        c.alignment = CENTER

        current_row = summary_row + 1

    print(f'    写入完成: {current_row - 2} 行 (含间隔和汇总行)')


def main():
    wb = openpyxl.load_workbook(DST)

    for name in SHEETS:
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}]')
        print(f'{"─"*50}')
        process_sheet(wb[name], name)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'步骤四完成 → {DST}')


if __name__ == '__main__':
    main()
