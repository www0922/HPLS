"""
步骤五：Lane编号、合并单元格、框线、文库环化填充
===============================================
1. B表 A/B/C 工作表:
   - A列写入 Lane编号 (A1,A2.../B1,B2.../C1,C2...)
   - 每组合并 A列 单元格
   - A列~K列 标框线
2. 文库环化 工作表:
   - 每组一行: A=Lane编号, E=汇总D, H=组第一行G, I=组第一行H
   - 公式: D=C*0.66*1, F=D, G=F/B, M=22*L/F
   - A列~K列 标框线
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config import DST
SHEETS = ['A', 'B', 'C']
CENTER = Alignment(horizontal='center', vertical='center')
ALL_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# 列索引
COL_A = 1
COL_D = 4
COL_G = 7
COL_H = 8
COL_K = 11
COL_O = 15
COL_P = 16


def read_groups(ws):
    """读取工作表, 按空白行分隔组, 每组含数据行+汇总行"""
    max_col = ws.max_column
    groups = []
    current = []

    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in values):
            if current:
                groups.append(current)
                current = []
            continue
        current.append(row)

    if current:
        groups.append(current)
    return groups


def read_groups_by_summary(ws):
    """读取工作表(无空白行), 按汇总行识别组: 汇总行B列为空但D列有值"""
    groups = []
    current = []
    for row in range(2, ws.max_row + 1):
        b = ws.cell(row=row, column=2).value
        d = ws.cell(row=row, column=4).value
        if b is None and d is not None:
            current.append(row)
            groups.append(current)
            current = []
        else:
            current.append(row)
    if current:
        groups.append(current)
    return groups


def _is_direct(ws, data_rows):
    """判断组是否为直接环化: G/O列含华大 或 P=已磷酸化"""
    for r in data_rows:
        g = str(ws.cell(row=r, column=COL_G).value or '')
        o = str(ws.cell(row=r, column=COL_O).value or '')
        p = str(ws.cell(row=r, column=COL_P).value or '')
        if '华大' in g or '华大' in o or p == '已磷酸化':
            return True
    return False


def _alt_fill_plate(ws, data_rows):
    """板号交替填充: 不同板号前缀(下划线前)交替蓝色"""
    blue = PatternFill(start_color='3399FF', end_color='3399FF', fill_type='solid')
    # 收集板号前缀及首次出现顺序
    seen = []
    for row in data_rows:
        plate = str(ws.cell(row=row, column=COL_K).value or '').strip()
        if not plate:
            continue
        if '_' in plate:
            prefix = plate.split('_')[0]
        elif '-' in plate:
            prefix = plate.rsplit('-', 1)[0]
        else:
            prefix = plate
        if prefix not in seen:
            seen.append(prefix)
    if len(seen) < 2:
        return
    # 交替: 0=不填, 1=蓝色, 2=不填, 3=蓝色...
    fill_prefixes = {p for i, p in enumerate(seen) if i % 2 == 1}
    for row in data_rows:
        plate = str(ws.cell(row=row, column=COL_K).value or '').strip()
        if not plate:
            continue
        if '_' in plate:
            prefix = plate.split('_')[0]
        elif '-' in plate:
            prefix = plate.rsplit('-', 1)[0]
        else:
            prefix = plate
        if prefix in fill_prefixes:
            ws.cell(row=row, column=COL_K).fill = blue


def process_abc_sheet(ws, name):
    """处理 A/B/C 工作表: Lane编号、合并、框线"""
    groups = read_groups_by_summary(ws)
    max_col = ws.max_column

    lane_info = []  # [(lane_name, first_g_val, first_h_val, summary_d_val), ...]

    for gi, rows in enumerate(groups):
        lane_name = f'{name}{gi + 1}'
        first_row = rows[0]
        last_row = rows[-1]

        # 数据行范围 (不含汇总行)
        data_rows = rows[:-1] if len(rows) > 1 else rows
        data_first = data_rows[0]
        data_last = data_rows[-1]

        # 取组内第一行的 G列、H列值
        first_g = ws.cell(row=first_row, column=COL_G).value
        first_h = ws.cell(row=first_row, column=COL_H).value
        # 汇总D: 对数据行D列求和
        summary_d = sum(
            float(ws.cell(row=r, column=COL_D).value or 0)
            for r in data_rows
        )

        lane_info.append((lane_name, first_g, first_h, summary_d))

        # 写入 Lane编号到数据行 A列
        for row in data_rows:
            cell = ws.cell(row=row, column=COL_A)
            cell.value = lane_name
            cell.alignment = CENTER

        # 合并数据行 A列
        if len(data_rows) > 1:
            ws.merge_cells(
                start_row=data_first, start_column=COL_A,
                end_row=data_last, end_column=COL_A
            )

        # 仅数据行 A列~K列 标框线（汇总行不加）
        for row in data_rows:
            for col in range(COL_A, COL_K + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = ALL_BORDER

        # 直接环化组: A列红色
        if _is_direct(ws, data_rows):
            red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            for row in data_rows:
                ws.cell(row=row, column=COL_A).fill = red

        # 板号交替填充: 不同板号前缀交替蓝色
        _alt_fill_plate(ws, data_rows)

        # 汇总行不加框线、不合并、不写Lane

    print(f'  Sheet {name}: {len(groups)} 组, Lane={name}1~{name}{len(groups)}')
    return lane_info


def fill_huanhua_sheet(ws_hh, all_lane_info):
    """填充文库环化工作表"""
    # 清空数据行（表头行1不动, clear_pooling已保留原始表头）
    if ws_hh.max_row > 1:
        ws_hh.delete_rows(2, ws_hh.max_row - 1)

    row = 2
    prev_letter = None

    for i, (lane, g_val, h_val, d_val) in enumerate(all_lane_info):
        cur_letter = lane[0] if lane else ''
        if prev_letter and cur_letter != prev_letter:
            row += 1
        prev_letter = cur_letter

        # A列: Lane编号
        c = ws_hh.cell(row=row, column=1)
        c.value = lane
        c.alignment = CENTER

        # B列, C列, L列: 留空(用户手工填)

        # D列: =C*0.66*1
        c = ws_hh.cell(row=row, column=4)
        c.value = f'=C{row}*0.66*1'
        c.alignment = CENTER

        # E列: 数据量
        c = ws_hh.cell(row=row, column=5)
        c.value = d_val
        c.alignment = CENTER

        # F列: =D
        c = ws_hh.cell(row=row, column=6)
        c.value = f'=D{row}'
        c.alignment = CENTER

        # G列: =F/B
        c = ws_hh.cell(row=row, column=7)
        c.value = f'=F{row}/B{row}'
        c.alignment = CENTER

        # H列: 文库类型
        c = ws_hh.cell(row=row, column=8)
        c.value = g_val
        c.alignment = CENTER

        # I列: 客户单位
        c = ws_hh.cell(row=row, column=9)
        c.value = h_val
        c.alignment = CENTER

        # J列(10): =40-G
        c = ws_hh.cell(row=row, column=10)
        c.value = f'=40-G{row}'
        c.alignment = CENTER

        # M列(13): =22*L/F
        c = ws_hh.cell(row=row, column=13)
        c.value = f'=22*L{row}/F{row}'
        c.alignment = CENTER

        # 框线 A~K
        for col in range(1, 12):
            ws_hh.cell(row=row, column=col).border = ALL_BORDER

        row += 1

    print(f'  文库环化: 写入 {len(all_lane_info)} 行')


def main():
    wb = openpyxl.load_workbook(DST)

    all_lane_info = []

    for name in SHEETS:
        print(f'\n{"─"*50}')
        print(f'处理工作表 [{name}]')
        print(f'{"─"*50}')
        info = process_abc_sheet(wb[name], name)
        all_lane_info.extend(info)

    print(f'\n{"─"*50}')
    print(f'填充 [文库环化]')
    print(f'{"─"*50}')
    fill_huanhua_sheet(wb['文库环化'], all_lane_info)

    wb.save(DST)
    print(f'\n{"="*50}')
    print(f'✅ 步骤五完成 → {DST}')


if __name__ == '__main__':
    main()
