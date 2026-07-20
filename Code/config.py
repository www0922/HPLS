"""
配置文件: 自动发现源表、管理路径
=================================
data/ 目录结构:
  - pool模板表.xlsx        : 输出模板
  - *上机*.xlsx            : A表 (上机表)
  - *质检*.xlsx            : C表 (外包文库质检总表)
  - *自建库*.xlsx          : D表 (自建库出库报告)

A表 sheet 名映射: 数字→英文字母 (1→A, 2→B, 3→C...)
"""
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / 'data'
DST = str(DATA / 'pool模板表.xlsx')


def find_file(keyword):
    """在 data/ 目录搜索包含关键词的文件, 返回绝对路径"""
    matches = list(DATA.glob(f'*{keyword}*.xlsx'))
    if not matches:
        raise FileNotFoundError(f'data/ 中没有包含 "{keyword}" 的 xlsx 文件')
    return str(matches[0])


def get_src_a():
    """A表: 上机表"""
    return find_file('上机')


def get_src_c():
    """C表: 外包文库质检总表"""
    return find_file('质检')


def get_src_d():
    """D表: 自建库出库报告"""
    return find_file('自建库')


def sheet_map_a():
    """扫描A表sheet名: 1→A, 2→B, 3→C..."""
    import openpyxl
    wb = openpyxl.load_workbook(get_src_a(), data_only=True)
    mapping = []
    for sn in wb.sheetnames:
        try:
            num = int(sn)
            letter = chr(ord('A') + num - 1)
            mapping.append((sn, letter))
        except ValueError:
            pass  # 跳过非数字sheet名
    wb.close()
    if not mapping:
        raise ValueError(f'A表中没有找到数字命名的sheet: {wb.sheetnames}')
    return mapping
