"""
统一格式: 字体、行高、数字格式、列宽
"""
import openpyxl
from openpyxl.styles import Font
from config import DST
FONT = Font(name='Times New Roman', size=10)

wb = openpyxl.load_workbook(DST)

for sn in wb.sheetnames:
    ws = wb[sn]

    # 表头行高
    if sn in ('A', 'B', 'C'):
        ws.row_dimensions[1].height = 90
    elif sn in ('文库环化', 'T7+制备'):
        ws.row_dimensions[1].height = 40

    # 数据行行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30

    # 字体 + 数字格式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            try:
                cell.font = FONT
                col_letter = cell.column_letter
                if col_letter == 'D':
                    cell.number_format = '0.00'
                elif col_letter in ('E', 'F', 'I'):
                    cell.number_format = '0.000'
            except AttributeError:
                pass  # 跳过MergedCell

    # 自动列宽
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            try:
                val = str(ws.cell(row=row, column=col_idx).value or '')
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
            except (AttributeError, TypeError):
                pass
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)

    print(f'{sn}: 格式已设置')

wb.save(DST)
print(f'\n✅ 完成 → {DST}')
